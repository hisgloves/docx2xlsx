import os
import docx
from docx import Document
import nltk
from openpyxl import Workbook, load_workbook

# Download necessary resources for sentence splitting
# nltk.download('punkt')

def extract_text_from_docx(docx_path):
    # Extracts text from a DOCX file
    doc = Document(docx_path)
    full_text = []
    for para in doc.paragraphs:
        full_text.append(para.text)
    return '\n'.join(full_text)

def split_text_into_sentences_and_lines(text):
    # Splits text into sentences and lines
    lines = text.split('\n')
    sentences = []
    for line in lines:
        if line.strip():  # Ensure we ignore empty lines
            # Split each line into sentences
            sentences.extend(nltk.tokenize.sent_tokenize(line))
    return sentences

def create_translation_excel(excel_path, sentences):
    # Creates an Excel file with two columns: 'Original' and 'Translation'
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Translation"
    sheet.append(['Original', 'Translation'])
    
    for sentence in sentences:
        if sentence.strip():  # Only add non-empty sentences
            sheet.append([sentence.strip(), ''])
    
    workbook.save(excel_path)

def count_characters_in_excel(excel_path):
    # Count characters in the 'Original' column of the Excel file excluding the header
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    char_count = 0
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        char_count += len(row[0])
    return char_count

def filter_text(text):
    # Remove bullets, tabs, item numbers, empty lines, and underscore separators
    filtered_lines = []
    for line in text.split('\n'):
        if line.strip() and not line.strip().startswith('________________________________________'):  # Ignore empty lines and separators
            # Remove leading item numbers or bullets (e.g., "1. ", "• ", "◦ ", "- ", "\t")
            filtered_line = line.lstrip("0123456789•◦- \t")
            filtered_lines.append(filtered_line)
    return ''.join(filtered_lines)

def count_characters_in_text(text):
    # Count characters in the given filtered text
    filtered_text = filter_text(text)
    return len(filtered_text)

def process_all_docx_files():
    # Processes all DOCX files in the current folder
    current_folder = os.getcwd()
    for filename in os.listdir(current_folder):
        if filename.endswith('.docx'):
            original_docx_path = os.path.join(current_folder, filename)
            new_excel_path = os.path.join(current_folder, f"{os.path.splitext(filename)[0]}_translate.xlsx")
            
            # Check if the corresponding Excel file already exists
            if os.path.exists(new_excel_path):
                print(f"Skipping {filename}, corresponding XLSX file already exists.")
                continue
            
            # Extract text from the original DOCX file
            text = extract_text_from_docx(original_docx_path)
            
            # Split the text into sentences and lines
            sentences = split_text_into_sentences_and_lines(text)
            
            # Create a new Excel file with a translation table
            create_translation_excel(new_excel_path, sentences)
            print(f"Processed {filename} and created {new_excel_path}")
            
            # Count characters in the original text and the generated Excel file
            # original_char_count = count_characters_in_text(text)
            # excel_char_count = count_characters_in_excel(new_excel_path)
            
            # Notify if the character counts do not match
            # if original_char_count != excel_char_count:
                # print(f"Warning: Character count mismatch for {filename}. DOCX: {original_char_count}, XLSX: {excel_char_count}")
            # else:
                # print(f"Character count matches for {filename}. DOCX: {original_char_count}, XLSX: {excel_char_count}")

# Run the script to process all DOCX files in the current folder
process_all_docx_files()
